import ee
import fiona
import numpy as np
import csv

print('Connecting...')
service = 'geo-image@vasapozac-geo-dip.iam.gserviceaccount.com'
path = '/home/vasilzach/geo_key.json'
credentials = ee.ServiceAccountCredentials(service, path)
ee.Initialize(credentials)
print('Connected')

features = []
tmp_name = 0
tmp_num = 0
dict_list = []
table_of_content = np.empty((67138,3), dtype=int)
i = 0 
names_ids = []

results = []
with open("EUForestspecies.csv") as csvfile:
    reader = csv.reader(csvfile) # change contents to floats
    for row in reader: # each row is a list
        results.append(row[3])
total_species = list(sorted(set(results)))

with fiona.open("/home/vasilzach/Data/Finland_Points.shp") as shapefile:
    for record in shapefile:
        name = record['properties']['Name']
        id = total_species.index(name)
        table_of_content[i,0] = i+1
        table_of_content[i,1] = id
        table_of_content[i,2] = int(record['properties']['fid_3'])
        # table_of_content[i,2] = int(record['properties']['CELLCODE'][5:8] + record['properties']['CELLCODE'][9:12])
        if tmp_name == 0:
            tmp_name = name
        if tmp_name != name:
            species = {
                "name": str(name),
                "number": tmp_num, 
            }
            dict_list.append(species)
            names_ids.append(str(name))
            tmp_name = name
            tmp_num = 0
        i = i+1
        geometry = record['geometry']['coordinates'][:]
        feature = ee.FeatureCollection(ee.Geometry.MultiPoint(geometry)).geometry().coordinates()
        features.append(feature)
        tmp_num = tmp_num + 1

print(len(features))

print(table_of_content)

import csv
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# filename = 'data_Est.xlsx'
wb = Workbook()
ws = wb.active

# script_directory = os.path.dirname(os.path.abspath(__file__))
# file_path = os.path.join(script_directory, filename)

# for row_idx, row_data in enumerate(table_of_content, start=1):
#     for col_idx, value in enumerate(row_data, start=1):
#         ws.cell(row=row_idx, column=col_idx, value=value)

# # Save the workbook
# wb.save(filename)

# # filename = 'species.xlsx'

# script_directory = os.path.dirname(os.path.abspath(__file__))
# file_path = os.path.join(script_directory, filename)

# for row_idx, row_data in enumerate(table_of_content, start=1):
#     for col_idx, value in enumerate(row_data, start=1):
#         ws.cell(row=row_idx, column=col_idx, value=value)

# # Save the workbook
# wb.save(filename)

species = {
    "name": name,
    "number": tmp_num, 
}
dict_list.append(species)
names_ids.append(str(name))

# print(names_ids)
# print(len(names_ids))
# print(dict_list)


newFs, counts = np.unique(table_of_content[:,2], return_counts=True)
print(len(newFs))


# multiple entries
Poi = []
Labels = []
indexes = np.where(counts > 1)[0]
elements = newFs[indexes]
print(elements.shape)
final_index = []
for element in elements:
    ilist = np.where(table_of_content[:,2] == element)[0]
    Poi.append([table_of_content[ilist,0], table_of_content[ilist,1], element])
    for i in ilist:
        final_index.append(i)
# print(final_index)
print(table_of_content[final_index,:].shape)
print(len(Poi))

# one entry
indexes = np.where(counts == 1)[0]
elements = newFs[indexes]
# print(elements)
for element in elements:
    ilist = np.where(table_of_content[:,2] == element)[0]
    Poi.append([table_of_content[ilist,0], table_of_content[ilist,1], element])
    for i in ilist:
        final_index.append(i)
print(table_of_content[final_index,:].shape)
print(len(Poi))

filename = 'merged_data_Fin.xlsx'
script_directory = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_directory, filename)

for row_idx, row in enumerate(Poi, start=1):
    for col_idx, value in enumerate(row, start=1):
        if isinstance(value, (np.ndarray, list)):
            value = ', '.join(str(num) for num in value)
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save(filename)